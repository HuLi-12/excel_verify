from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any

import openpyxl
import yaml
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class ColumnMeta:
    index: int
    excel_column: str
    name: str
    compare_key: str
    data_key: str
    category: str
    field: str
    sub_field: str
    raw_header_empty: bool


@dataclass(frozen=True)
class TableMeta:
    columns: list[ColumnMeta]
    data_start_row: int | None = None

    @property
    def field_names(self) -> list[str]:
        return [column.name for column in self.columns if column.name]

    @property
    def field_keys(self) -> list[str]:
        return [column.compare_key for column in self.columns if column.name]


def load_rules(rule_path: str | Path) -> dict[str, Any]:
    with Path(rule_path).open("r", encoding="utf-8") as file:
        return yaml.safe_load(file) or {}


def load_sheet_values(excel_path: str | Path, sheet_name: str | None = None) -> list[list[Any]]:
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name and sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    merged_values: dict[tuple[int, int], Any] = {}
    for merged_range in sheet.merged_cells.ranges:
        value = sheet.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, col)] = value

    rows: list[list[Any]] = []
    for row in range(1, sheet.max_row + 1):
        values: list[Any] = []
        for col in range(1, sheet.max_column + 1):
            values.append(merged_values.get((row, col), sheet.cell(row, col).value))
        rows.append(values)
    return rows


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return " ".join(text.split())


def build_column_name(category: str, field: str, sub_field: str) -> str:
    parts = [part for part in [category, field, sub_field] if part]
    cleaned: list[str] = []
    for part in parts:
        if not cleaned or cleaned[-1] != part:
            cleaned.append(part)
    return "_".join(cleaned)


def build_compare_key(category: str, field: str, sub_field: str) -> str:
    compare_sub_field = sub_field
    if (
        sub_field == field
        or (field == "序号" and sub_field.isdigit())
        or (field == "资产名称" and sub_field == "资产全称")
    ):
        compare_sub_field = ""
    return canonical_field_name(build_column_name(category, field, compare_sub_field))


def canonical_field_name(field_name: str) -> str:
    text = normalize_cell(field_name)
    translation = str.maketrans(
        {
            "（": "(",
            "）": ")",
            "，": ",",
            "。": ".",
            "／": "/",
        }
    )
    text = text.translate(translation)
    text = re.sub(r"\s+", "", text)
    return text


def parse_columns(rows: list[list[Any]], header_rows: dict[str, int]) -> list[ColumnMeta]:
    category_row = header_rows["category"] - 1
    field_row = header_rows["field"] - 1
    sub_field_row = header_rows["sub_field"] - 1
    max_columns = max((len(row) for row in rows), default=0)

    columns: list[ColumnMeta] = []
    seen: dict[str, int] = {}
    for index in range(max_columns):
        category = _cell(rows, category_row, index)
        field = _cell(rows, field_row, index)
        sub_field = _cell(rows, sub_field_row, index)
        raw_header_empty = not any([category, field, sub_field])
        name = "" if raw_header_empty else build_column_name(category, field, sub_field)
        compare_key = "" if raw_header_empty else build_compare_key(category, field, sub_field)
        data_key = name or f"__EMPTY_HEADER_{get_column_letter(index + 1)}"
        if data_key in seen:
            seen[data_key] += 1
            data_key = f"{data_key}__{get_column_letter(index + 1)}"
        else:
            seen[data_key] = 1
        columns.append(
            ColumnMeta(
                index=index,
                excel_column=get_column_letter(index + 1),
                name=name,
                compare_key=compare_key,
                data_key=data_key,
                category=category,
                field=field,
                sub_field=sub_field,
                raw_header_empty=raw_header_empty,
            )
        )
    return columns


def _cell(rows: list[list[Any]], row_index: int, col_index: int) -> str:
    if row_index < 0 or row_index >= len(rows):
        return ""
    row = rows[row_index]
    if col_index < 0 or col_index >= len(row):
        return ""
    return normalize_cell(row[col_index])
