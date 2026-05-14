from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill


ERROR_FILL = PatternFill(fill_type="solid", fgColor="FFFFC7CE")
WARNING_FILL = PatternFill(fill_type="solid", fgColor="FFFFEB9C")
NORMALIZED_FILL = PatternFill(fill_type="solid", fgColor="FFDDEBF7")
ERROR_FONT_COLOR = "FF9C0006"
DETAIL_SHEET_NAME = "错误明细"
DETAIL_HEADERS = ["Excel行号", "Excel列号", "字段名", "原始值", "错误类型", "错误说明"]
WARNING_SHEET_NAME = "警告明细"
WARNING_HEADERS = ["Excel行号", "Excel列号", "字段名", "原始值", "处理方式", "警告类型", "警告说明"]
NORMALIZED_SHEET_NAME = "自动修正明细"
NORMALIZED_HEADERS = ["Excel行号", "Excel列号", "字段名", "原始值", "修正后", "修正说明"]


def write_validated_workbook(
    source_workbook_path: str | Path,
    output_path: str | Path,
    errors: list[dict[str, Any]],
    replace_error_values: bool = True,
    generate_error_detail_sheet: bool = True,
) -> None:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = openpyxl.load_workbook(source_workbook_path)
    data_sheet = workbook.worksheets[0]

    for finding in errors:
        row = finding.get("Excel行号")
        column = finding.get("Excel列号")
        if not isinstance(row, int) or not isinstance(column, str) or not column:
            continue
        cell = data_sheet[f"{column}{row}"]
        _apply_finding_to_cell(cell, finding, replace_error_values)

    if generate_error_detail_sheet:
        _write_detail_sheets(workbook, errors)
    workbook.save(output)


def _apply_finding_to_cell(cell: openpyxl.cell.cell.Cell, finding: dict[str, Any], replace_error_values: bool) -> None:
    status = finding.get("status", "ERROR")
    if status == "NORMALIZED":
        cell.value = finding.get("输出值", cell.value)
        cell.fill = NORMALIZED_FILL
    elif status == "WARNING":
        cell.fill = WARNING_FILL
    else:
        if replace_error_values:
            cell.value = finding.get("输出值", "/")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = ERROR_FILL
        font = copy(cell.font)
        font.color = ERROR_FONT_COLOR
        cell.font = font
    cell.comment = Comment(_comment_text(finding), "asset-validator")


def _write_detail_sheets(workbook: openpyxl.Workbook, findings: list[dict[str, Any]]) -> None:
    error_findings = [finding for finding in findings if finding.get("status", "ERROR") == "ERROR"]
    warning_findings = [finding for finding in findings if finding.get("status") == "WARNING"]
    normalized_findings = [finding for finding in findings if finding.get("status") == "NORMALIZED"]

    _write_sheet(workbook, DETAIL_SHEET_NAME, DETAIL_HEADERS, error_findings)
    if warning_findings:
        _write_sheet(workbook, WARNING_SHEET_NAME, WARNING_HEADERS, [_warning_row(finding) for finding in warning_findings])
    if normalized_findings:
        _write_sheet(
            workbook,
            NORMALIZED_SHEET_NAME,
            NORMALIZED_HEADERS,
            [_normalized_row(finding) for finding in normalized_findings],
        )


def _write_sheet(
    workbook: openpyxl.Workbook,
    sheet_name: str,
    headers: list[str],
    rows: list[dict[str, Any]],
) -> None:
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(headers)
    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = header_font

    for row in rows:
        sheet.append([row.get(header, "") for header in headers])


def _warning_row(finding: dict[str, Any]) -> dict[str, Any]:
    return {
        **finding,
        "处理方式": "保留原值并标黄",
    }


def _normalized_row(finding: dict[str, Any]) -> dict[str, Any]:
    return {
        **finding,
        "修正后": finding.get("输出值", ""),
    }


def _comment_text(finding: dict[str, Any]) -> str:
    status = finding.get("status", "ERROR")
    if status == "NORMALIZED":
        return f"自动修正：{finding.get('修正说明', '')}"
    if status == "WARNING":
        return f"警告：{finding.get('警告说明', '')}"
    return f"{finding.get('错误类型', '')}：{finding.get('错误说明', '')}"
