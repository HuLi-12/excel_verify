from __future__ import annotations

from copy import copy
from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill


ERROR_FILL = PatternFill(fill_type="solid", fgColor="FFFFC7CE")
ERROR_FONT_COLOR = "FF9C0006"
DETAIL_SHEETS = ["汇总概览", "结构错误", "数据错误"]


def write_report(
    structure_errors: list[dict[str, Any]],
    data_errors: list[dict[str, Any]],
    output_path: str | Path,
    total_rows: int = 0,
    source_workbook_path: str | Path | None = None,
) -> None:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    data_error_counter = Counter(error["错误类型"] for error in data_errors)
    overview = pd.DataFrame(
        [
            {"指标": "总数据行数", "数量": total_rows},
            {"指标": "结构错误数", "数量": len(structure_errors)},
            {"指标": "数据错误数", "数量": len(data_errors)},
            {"指标": "必填项错误数", "数量": data_error_counter.get("必填项为空", 0)},
            {"指标": "数值格式错误数", "数量": data_error_counter.get("数值格式错误", 0)},
            {"指标": "枚举错误数", "数量": data_error_counter.get("枚举错误", 0)},
            {"指标": "手机号错误数", "数量": data_error_counter.get("手机号错误", 0)},
        ]
    )

    if source_workbook_path:
        workbook = openpyxl.load_workbook(source_workbook_path)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)

    _mark_error_cells(workbook, structure_errors + data_errors)
    _replace_detail_sheets(
        workbook,
        overview=overview,
        structure_errors=pd.DataFrame(structure_errors),
        data_errors=pd.DataFrame(data_errors),
    )
    workbook.save(output)


def _mark_error_cells(workbook: Workbook, errors: list[dict[str, Any]]) -> None:
    if not workbook.worksheets:
        return
    sheet = workbook.worksheets[0]
    grouped_errors: dict[str, list[dict[str, Any]]] = {}

    for error in errors:
        row = error.get("Excel行号")
        column = error.get("Excel列号")
        if not isinstance(row, int) or not isinstance(column, str) or not column:
            continue
        grouped_errors.setdefault(f"{column}{row}", []).append(error)

    for coordinate, cell_errors in grouped_errors.items():
        cell = sheet[coordinate]
        cell.fill = ERROR_FILL
        font = copy(cell.font)
        font.color = ERROR_FONT_COLOR
        cell.font = font
        cell.comment = Comment(_comment_text(cell_errors), "asset-validator")


def _comment_text(errors: list[dict[str, Any]]) -> str:
    lines = []
    for error in errors:
        error_type = error.get("错误类型", "")
        message = error.get("错误说明", "")
        field = error.get("字段名") or error.get("当前字段") or error.get("模板字段", "")
        lines.append(f"{error_type}：{field}，{message}")
    return "\n".join(lines)


def _replace_detail_sheets(
    workbook: Workbook,
    overview: pd.DataFrame,
    structure_errors: pd.DataFrame,
    data_errors: pd.DataFrame,
) -> None:
    for sheet_name in DETAIL_SHEETS:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    _append_dataframe_sheet(workbook, "汇总概览", overview)
    _append_dataframe_sheet(workbook, "结构错误", structure_errors)
    _append_dataframe_sheet(workbook, "数据错误", data_errors)


def _append_dataframe_sheet(workbook: Workbook, sheet_name: str, dataframe: pd.DataFrame) -> None:
    sheet = workbook.create_sheet(sheet_name)
    if dataframe.empty:
        return

    sheet.append(list(dataframe.columns))
    for _, row in dataframe.iterrows():
        sheet.append([row[column] for column in dataframe.columns])

    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = header_font
