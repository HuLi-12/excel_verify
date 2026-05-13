import openpyxl

from src.data_region import find_data_start_row
from src.excel_loader import load_sheet_values, parse_columns


def test_find_data_start_row_from_first_sequence_one(tmp_path):
    path = tmp_path / "data.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资产基本信息调研表"
    ws["A1"] = "汇总表"
    ws["A2"] = "一、资产基础法律属性"
    ws["A3"] = "序号"
    ws["A4"] = "序号"
    ws["A5"] = "示例"
    ws["A6"] = 1
    ws["A7"] = 2
    wb.save(path)

    rows = load_sheet_values(path, "资产基本信息调研表")
    columns = parse_columns(rows, {"category": 2, "field": 3, "sub_field": 4})

    assert find_data_start_row(rows, columns) == 6
