import openpyxl

from src.template_parser import parse_template_rules


def test_parse_template_rules_uses_sample_row_for_rule_inference(tmp_path):
    path = tmp_path / "template.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资产基本信息调研表"
    ws["A1"] = "一、资产基础法律属性"
    ws["A2"] = "资产来源"
    ws["A3"] = "划转/购入/自建"
    ws["A4"] = "划转/购入/自建"
    wb.save(path)

    rules = parse_template_rules(path, "资产基本信息调研表")

    assert rules[0].field_name == "一、资产基础法律属性_资产来源_划转/购入/自建"
    assert rules[0].excel_column == "A"
    assert rules[0].sample_value == "划转/购入/自建"
    assert rules[0].rule["type"] == "enum"
