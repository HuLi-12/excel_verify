import openpyxl

from src.template_driven_validator import validate_workbook


def test_template_driven_pipeline_starts_at_sequence_one_and_returns_cell_errors(tmp_path):
    template = tmp_path / "template.xlsx"
    data = tmp_path / "data.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资产基本信息调研表"
    ws["A1"] = "一、资产基础法律属性"
    ws["A2"] = "序号"
    ws["A3"] = "序号"
    ws["A4"] = "1"
    ws["B1"] = "一、资产基础法律属性"
    ws["B2"] = "资产来源"
    ws["B3"] = "划转/购入/自建"
    ws["B4"] = "划转/购入/自建"
    wb.save(template)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资产基本信息调研表"
    ws["A1"] = "汇总"
    ws["A2"] = "一、资产基础法律属性"
    ws["A3"] = "序号"
    ws["A4"] = "序号"
    ws["A5"] = "说明行"
    ws["A6"] = 1
    ws["B2"] = "一、资产基础法律属性"
    ws["B3"] = "资产来源"
    ws["B4"] = "划转/购入/自建"
    ws["B6"] = "新建"
    wb.save(data)

    result = validate_workbook(template, data, "资产基本信息调研表")

    assert result.data_start_row == 6
    assert len(result.errors) == 1
    assert result.errors[0]["Excel行号"] == 6
    assert result.errors[0]["Excel列号"] == "B"
    assert result.errors[0]["原始值"] == "新建"
