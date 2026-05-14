import openpyxl

from src.output_writer import write_validated_workbook


def test_output_writer_replaces_error_cell_with_slash_and_keeps_original_in_detail(tmp_path):
    source = tmp_path / "data.xlsx"
    output = tmp_path / "output.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资产基本信息调研表"
    ws["B6"] = "新建"
    wb.save(source)

    write_validated_workbook(
        source,
        output,
        [
            {
                "Excel行号": 6,
                "Excel列号": "B",
                "字段名": "资产来源",
                "原始值": "新建",
                "错误类型": "枚举错误",
                "错误说明": "允许值：划转/购入/自建",
            }
        ],
    )

    result = openpyxl.load_workbook(output)
    sheet = result["资产基本信息调研表"]
    assert sheet["B6"].value == "/"
    assert sheet["B6"].fill.fgColor.rgb == "FFFFC7CE"
    assert sheet["B6"].alignment.horizontal == "center"
    assert sheet["B6"].alignment.vertical == "center"
    detail = result["错误明细"]
    assert detail["D2"].value == "新建"


def test_output_writer_can_highlight_without_replacing_value(tmp_path):
    source = tmp_path / "data.xlsx"
    output = tmp_path / "output.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资产基本信息调研表"
    ws["B6"] = "新建"
    wb.save(source)

    write_validated_workbook(
        source,
        output,
        [
            {
                "Excel行号": 6,
                "Excel列号": "B",
                "字段名": "资产来源",
                "原始值": "新建",
                "错误类型": "枚举错误",
                "错误说明": "允许值：划转/购入/自建",
            }
        ],
        replace_error_values=False,
    )

    result = openpyxl.load_workbook(output)
    sheet = result["资产基本信息调研表"]
    assert sheet["B6"].value == "新建"
    assert sheet["B6"].fill.fgColor.rgb == "FFFFC7CE"


def test_output_writer_can_skip_error_detail_sheet(tmp_path):
    source = tmp_path / "data.xlsx"
    output = tmp_path / "output.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资产基本信息调研表"
    ws["B6"] = "新建"
    wb.save(source)

    write_validated_workbook(
        source,
        output,
        [
            {
                "Excel行号": 6,
                "Excel列号": "B",
                "字段名": "资产来源",
                "原始值": "新建",
                "错误类型": "枚举错误",
                "错误说明": "允许值：划转/购入/自建",
            }
        ],
        generate_error_detail_sheet=False,
    )

    result = openpyxl.load_workbook(output)
    assert "错误明细" not in result.sheetnames


def test_output_writer_handles_normalized_warning_and_error_details(tmp_path):
    source = tmp_path / "data.xlsx"
    output = tmp_path / "output.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资产基本信息调研表"
    ws["B6"] = "500;1200"
    ws["C6"] = "无"
    ws["D6"] = "新建"
    wb.save(source)

    write_validated_workbook(
        source,
        output,
        [
            {
                "status": "NORMALIZED",
                "Excel行号": 6,
                "Excel列号": "B",
                "字段名": "距离列表",
                "原始值": "500;1200",
                "输出值": "500；1200",
                "修正说明": "英文分号已统一为中文分号",
            },
            {
                "status": "WARNING",
                "Excel行号": 6,
                "Excel列号": "C",
                "字段名": "建筑面积",
                "原始值": "无",
                "输出值": "无",
                "警告类型": "结构化字段填写无",
                "警告说明": "该字段应填写结构化数据，当前填写“无”，请人工确认",
            },
            {
                "status": "ERROR",
                "Excel行号": 6,
                "Excel列号": "D",
                "字段名": "资产来源",
                "原始值": "新建",
                "输出值": "/",
                "错误类型": "枚举错误",
                "错误说明": "允许值：划转/购入/自建",
            },
        ],
    )

    result = openpyxl.load_workbook(output)
    sheet = result["资产基本信息调研表"]
    assert sheet["B6"].value == "500；1200"
    assert sheet["C6"].value == "无"
    assert sheet["C6"].fill.fgColor.rgb == "FFFFEB9C"
    assert sheet["D6"].value == "/"
    assert result["自动修正明细"]["D2"].value == "500;1200"
    assert result["警告明细"]["D2"].value == "无"
    assert result["错误明细"]["D2"].value == "新建"
