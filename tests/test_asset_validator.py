from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from src.summary_parser import parse_summary
from src.template_parser import parse_template
from src.validators import validate_data, validate_structure
from src.report_writer import write_report


def _write_rules(path: Path) -> Path:
    path.write_text(
        """
sheet_name: 资产基本信息调研表
template:
  header_rows:
    category: 1
    field: 2
    sub_field: 3
    sample: 4
    note: 5
summary:
  title_row: 1
  header_rows:
    category: 2
    field: 3
    sub_field: 4
  data_start_row: 5
special_empty_values:
  - ""
  - "/"
  - "无"
  - "暂无"
  - "未提供"
  - "未挂表"
  - "在建"
  - "nan"
required_keywords:
  - 资产名称
  - 手机
strict_numeric_keywords:
  - 面积
numeric_with_unit_keywords:
  - 距离
  - 单价
comparison_numeric_keywords:
  - 规划上限
  - 下限
range_or_description_keywords:
  - 所在层数
  - 总层数
free_text_numeric_keywords:
  - 免租期意向
numeric_special_values:
  - 面议
  - 永久
numeric_status_suffix_values:
  - 在建
  - 待建
enum_rules:
  资产来源:
    - 划转
    - 购入
    - 自建
    - 代管
    - 安置房
    - 国家财政资金
    - 划拨
phone_rules:
  - 手机
""",
        encoding="utf-8",
    )
    return path


def _save_template(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资产基本信息调研表"
    ws.merge_cells("A1:D1")
    ws["A1"] = "一、资产基础法律属性"
    ws["A2"] = "序号"
    ws["B2"] = "资产名称"
    ws["B3"] = "资产全称"
    ws["C2"] = "资产来源"
    ws["D2"] = "有产权面积"
    ws["D3"] = "建筑面积（平方米）"
    ws.merge_cells("E1:F1")
    ws["E1"] = "六、联系人"
    ws["E2"] = "联系人姓名"
    ws["F2"] = "手机"
    wb.save(path)
    return path


def _save_summary(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资产基本信息调研表"
    ws["A1"] = "汇总表"
    ws.merge_cells("A2:D2")
    ws["A2"] = "一、资产基础法律属性"
    ws["A3"] = "序号"
    ws["B3"] = "资产名称"
    ws["B4"] = "资产全称"
    ws["C3"] = "资产来源"
    ws["D3"] = "有产权面积"
    ws["D4"] = "建筑面积（平方米）"
    ws.merge_cells("E2:F2")
    ws["E2"] = "六、联系人"
    ws["E3"] = "联系人姓名"
    ws["F3"] = "手机"
    ws["G5"] = "城建集团"
    ws.append([1, "项目A", "新建", "约3000㎡", "张三", "123456", "一建公司"])
    ws.append([2, "/", "划转", "1200.5", "李四", "13800138000", None])
    wb.save(path)
    return path


def test_parse_template_builds_stable_unique_field_names(tmp_path):
    rules = _write_rules(tmp_path / "rules.yaml")
    template = _save_template(tmp_path / "template.xlsx")

    meta = parse_template(template, rules)

    assert [col.name for col in meta.columns] == [
        "一、资产基础法律属性_序号",
        "一、资产基础法律属性_资产名称_资产全称",
        "一、资产基础法律属性_资产来源",
        "一、资产基础法律属性_有产权面积_建筑面积（平方米）",
        "六、联系人_联系人姓名",
        "六、联系人_手机",
    ]


def test_structure_validation_reports_blank_header_column_with_data(tmp_path):
    rules = _write_rules(tmp_path / "rules.yaml")
    template = _save_template(tmp_path / "template.xlsx")
    summary = _save_summary(tmp_path / "summary.xlsx")

    template_meta = parse_template(template, rules)
    summary_df, summary_meta = parse_summary(summary, rules)
    errors = validate_structure(template_meta, summary_meta, summary_df)

    assert any(error["错误类型"] == "空表头列有数据" for error in errors)
    blank_error = next(error for error in errors if error["错误类型"] == "空表头列有数据")
    assert blank_error["Excel列号"] == "G"
    assert blank_error["错误值"] == "城建集团"


def test_structure_validation_ignores_cosmetic_header_differences(tmp_path):
    rules = _write_rules(tmp_path / "rules.yaml")
    template_wb = openpyxl.Workbook()
    template_ws = template_wb.active
    template_ws.title = "资产基本信息调研表"
    template_ws["A1"] = "一、资产基础法律属性"
    template_ws["A2"] = "序号"
    template_ws["A3"] = "1"
    template_ws["B1"] = "二、空间与物理指标"
    template_ws["B2"] = "有产权面积"
    template_ws["B3"] = "分摊土地使用权面积（平方米）"
    template_ws["C1"] = "一、资产基础法律属性"
    template_ws["C2"] = "资产名称"
    template_ws["C3"] = "资产全称"
    template_path = tmp_path / "template.xlsx"
    template_wb.save(template_path)

    summary_wb = openpyxl.Workbook()
    summary_ws = summary_wb.active
    summary_ws.title = "资产基本信息调研表"
    summary_ws["A1"] = "汇总表"
    summary_ws["A2"] = "一、资产基础法律属性"
    summary_ws["A3"] = "序号"
    summary_ws["A4"] = "序号"
    summary_ws["B2"] = "二、空间与物理指标"
    summary_ws["B3"] = "有产权面积"
    summary_ws["B4"] = "分摊土地使用权面积 （平方米）"
    summary_ws["C2"] = "一、资产基础法律属性"
    summary_ws["C3"] = "资产名称"
    summary_ws["C4"] = "资产名称"
    summary_path = tmp_path / "summary.xlsx"
    summary_wb.save(summary_path)

    template_meta = parse_template(template_path, rules)
    summary_df, summary_meta = parse_summary(summary_path, rules)
    errors = validate_structure(template_meta, summary_meta, summary_df)

    assert errors == []


def test_data_validation_reports_required_numeric_enum_and_phone_errors(tmp_path):
    rules = _write_rules(tmp_path / "rules.yaml")
    summary = _save_summary(tmp_path / "summary.xlsx")

    summary_df, summary_meta = parse_summary(summary, rules)
    errors = validate_data(summary_df, summary_meta, rules)
    error_types = {error["错误类型"] for error in errors}

    assert "必填项为空" in error_types
    assert "数值格式错误" in error_types
    assert "枚举错误" in error_types
    assert "手机号错误" in error_types

    required = next(error for error in errors if error["错误类型"] == "必填项为空")
    assert required["Excel行号"] == 7
    assert "资产名称" in required["字段名"]

    numeric = next(error for error in errors if error["错误类型"] == "数值格式错误")
    assert numeric["错误值"] == "约3000㎡"

    phone = next(error for error in errors if error["错误类型"] == "手机号错误")
    assert phone["Excel列号"] == "F"


def test_data_validation_accepts_excel_number_phone_and_expanded_asset_sources(tmp_path):
    rules = _write_rules(tmp_path / "rules.yaml")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资产基本信息调研表"
    ws["A1"] = "汇总表"
    ws["A2"] = "一、资产基础法律属性"
    ws["A3"] = "资产来源"
    ws["A4"] = "划转/购入/自建"
    ws["B2"] = "六、联系人"
    ws["B3"] = "手机"
    ws.append(["代管", 13979135505.0])
    ws.append(["安置房", "18970092883.0"])
    ws.append(["国家财政资金", 15279141997])
    ws.append(["划拨", "13800138000"])
    summary = tmp_path / "summary.xlsx"
    wb.save(summary)

    summary_df, summary_meta = parse_summary(summary, rules)
    errors = validate_data(summary_df, summary_meta, rules)

    assert errors == []


def test_data_validation_uses_more_precise_numeric_modes(tmp_path):
    rules = _write_rules(tmp_path / "rules.yaml")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资产基本信息调研表"
    ws["A1"] = "汇总表"
    ws["A2"] = "二、空间与物理指标"
    ws["A3"] = "有产权面积"
    ws["A4"] = "建筑面积（平方米）"
    ws["B2"] = "四、配套设施与周边环境"
    ws["B3"] = "交通通达度"
    ws["B4"] = "附近地铁站距离（m）"
    ws["C2"] = "二、空间与物理指标"
    ws["C3"] = "建筑层数与层高"
    ws["C4"] = "所在层数（F）"
    ws["D2"] = "五、招租意向与商务条件"
    ws["D3"] = "租金参考价"
    ws["D4"] = "单价（元/㎡/月）"
    ws["E2"] = "三、土地经济技术指标"
    ws["E3"] = "容积率"
    ws["E4"] = "规划上限"
    ws["F2"] = "五、招租意向与商务条件"
    ws["F3"] = "免租期意向"
    ws["F4"] = "装修/改建免租期（月）"
    ws.append(["约3000㎡", "15公里", "1——7", "面议", "≤2.5", "看承租方装修方案后确定"])
    ws.append(["51243.96 （在建）", "6公里", "1至2层", "120", "≥30%", "面议"])
    summary = tmp_path / "summary.xlsx"
    wb.save(summary)

    summary_df, summary_meta = parse_summary(summary, rules)
    errors = validate_data(summary_df, summary_meta, rules)

    assert len(errors) == 1
    assert errors[0]["字段名"] == "二、空间与物理指标_有产权面积_建筑面积（平方米）"
    assert errors[0]["错误值"] == "约3000㎡"


def test_write_report_copies_summary_workbook_and_marks_error_cells_red(tmp_path):
    summary = _save_summary(tmp_path / "summary.xlsx")
    output = tmp_path / "report.xlsx"
    errors = [
        {
            "错误类型": "手机号错误",
            "Excel行号": 6,
            "Excel列号": "F",
            "字段名": "六、联系人_手机",
            "错误值": "123456",
            "错误说明": "手机号格式不正确",
        }
    ]

    write_report(
        structure_errors=[],
        data_errors=errors,
        output_path=output,
        total_rows=2,
        source_workbook_path=summary,
    )

    workbook = openpyxl.load_workbook(output)
    sheet = workbook["资产基本信息调研表"]

    assert sheet["A1"].value == "汇总表"
    assert "A2:D2" in [str(cell_range) for cell_range in sheet.merged_cells.ranges]
    assert sheet["F6"].fill.fgColor.rgb == "FFFFC7CE"
    assert sheet["F6"].font.color.rgb == "FF9C0006"
    assert sheet["F6"].comment is not None
    assert "手机号错误" in sheet["F6"].comment.text
    assert "汇总概览" in workbook.sheetnames
    assert "数据错误" in workbook.sheetnames
